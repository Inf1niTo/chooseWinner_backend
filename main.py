import pandas as pd
import math
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
CORS(app)

# MongoDB connection settings
MONGO_URI = "mongodb://localhost:27018/"  # Подставьте свой URI
DATABASE_NAME = "lottery_db"
BUYERS_COLLECTION = "buyers"
WINNERS_COLLECTION = "winners"

# Подключение к MongoDB
client = MongoClient(MONGO_URI)
db = client[DATABASE_NAME]
buyers_collection = db[BUYERS_COLLECTION]
winners_collection = db[WINNERS_COLLECTION]


@app.route('/api/upload', methods=['POST'])
def upload_participants():
    """
    API-маршрут для загрузки CSV файла с участниками.

    Ожидает файл в поле 'file' в POST-запросе.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Файл не найден'}), 400

        file = request.files['file']
        print(f"Имя файла: {file.filename}")  # Отладочный вывод

        if file.filename == '':
            return jsonify({'error': 'Файл не выбран'}), 400

        try:
            # --- Чтение данных из CSV файла ---
            all_buyers_df = pd.read_csv(file, header=None, names=['Data'], skiprows=1)
        except Exception as e:
            return jsonify({'error': f"Ошибка при чтении CSV: {str(e)}"}), 500

        # --- Извлечение ID и данных ---

        all_buyers_df['Digital_ID'] = range(1, len(all_buyers_df) + 1)
        all_buyers_df['Email'] = all_buyers_df['Data'].str.split(',').str[0]
        all_buyers_df = all_buyers_df.drop('Data', axis=1)

        # --- Перенос данных участников в MongoDB ---

        buyers_data = all_buyers_df.to_dict("records")
        buyers_collection.delete_many({})  # Очищаем коллекцию перед добавлением
        buyers_collection.insert_many(buyers_data)

        return jsonify({'message': 'Данные успешно загружены'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/calculate_winners', methods=['POST'])
def calculate_winners():
    """API-маршрут для расчета победителей."""
    try:
        data = request.get_json()
        num_iterations = int(data['num_winners'])
        draw_number = int(data['draw_number'])
        draw_category = data.get('draw_category', '')

        # --- Определение коллекции победителей ---
        winners_collection_name = f"winners_{draw_number}"
        winners_collection = db[winners_collection_name]

        # --- Получение ID существующих победителей ---
        existing_winners = winners_collection.distinct("Digital_ID")

        # --- Расчет параметров ---
        num_buers = buyers_collection.count_documents({})

        if num_buers - len(existing_winners) < num_iterations:
            return jsonify({
                'error': 'Недостаточно участников для выбранного количества победителей.'
            }), 400

        # --- Получение курса валюты ---
        url1 = 'http://www.cbr.ru/scripts/XML_daily.asp'
        daily = pd.read_xml(url1, encoding='cp1251')
        course = daily.iloc[3]['Value'].replace(',', '.')
        Cf = float(course[-4:-2])

        if Cf == 00:
            Cf = 1

        # --- Расчет шага выбора победителя ---
        step = math.ceil((num_buers - len(existing_winners)) / Cf)

        # --- Выбор победителей ---
        def choose_winner(step, total_tickets, winners):
            """Функция для определения победителей."""
            new_winners = []
            current_ticket = step
            checked_tickets = set(winners)

            while len(new_winners) < num_iterations and len(checked_tickets) < total_tickets:
                if current_ticket > total_tickets:
                    current_ticket = (current_ticket % total_tickets) + 1

                while current_ticket in checked_tickets:
                    current_ticket += 1
                    if current_ticket > total_tickets:
                        current_ticket = (current_ticket % total_tickets) + 1

                new_winners.append(current_ticket)
                checked_tickets.add(current_ticket)
                current_ticket += step

            return new_winners

        # Выбор новых победителей
        new_winners = choose_winner(step, num_buers, existing_winners)

        # --- Сохранение новых победителей ---
        new_winners_data = list(buyers_collection.find(
            {'Digital_ID': {'$in': new_winners}},
            {'_id': 0, 'Digital_ID': 1, 'Email': 1, 'draw_category': 1}
        ))

        for winner in new_winners_data:
            winner['draw_category'] = draw_category

        winners_collection.insert_many(new_winners_data)

        return jsonify({
            'message': 'Победители успешно рассчитаны',
            'total_participants': num_buers,
            'selected_winners': num_iterations,
            'course': course,
            'step': step
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/winners/<int:draw_number>', methods=['GET'])
def get_winners(draw_number):

    """API-маршрут для получения списка НОВЫХ победителей по номеру розыгрыша."""

    # --- Определение коллекции победителей ---
    winners_collection_name = f"winners_{draw_number}"
    winners_collection = db[winners_collection_name]

    # --- Определение количества новых победителей для вывода ---
    last_winners_count = int(request.args.get('count', 0))

    if (last_winners_count > 0):
        winners = list(winners_collection.find().sort('_id', -1).limit(last_winners_count))
    else:
        winners = []

    # Исключаем _id из ответа
    winners = [{k: v for k, v in winner.items() if k != '_id'} for winner in winners]
    return jsonify(winners), 200

@app.route('/api/winners/all/<int:draw_number>', methods=['GET'])
def get_all_winners(draw_number):
    """API-маршрут для получения списка ВСЕХ победителей по номеру розыгрыша."""

    # --- Определение коллекции победителей ---
    winners_collection_name = f"winners_{draw_number}"
    winners_collection = db[winners_collection_name]

    winners = list(winners_collection.find({}, {'_id': 0}))
    return jsonify(winners), 200


@app.route('/api/generate_report', methods=['POST'])
def generate_report():
    """API-маршрут для генерации Excel-отчета."""
    try:
        data = request.get_json()

        # --- Создание Excel-файла ---
        wb = Workbook()

        # --- Лист 1:  Информация о розыгрыше ---
        ws1 = wb.active
        ws1.title = "Информация"
        for row in data['Sheet1']:
            ws1.append([row['Название'], row['Значение']])

        # --- Лист 2:  Текущие победители ---
        ws2 = wb.create_sheet(title="Текущие победители")
        for r in dataframe_to_rows(pd.DataFrame(data['Текущие победители']), index=False, header=True):
            ws2.append(r)

        # --- Лист 3:  Все победители ---
        ws3 = wb.create_sheet(title="Все победители")
        for r in dataframe_to_rows(pd.DataFrame(data['Все победители']), index=False, header=True):
            ws3.append(r)

        # --- Сохранение Excel-файла в байтовый поток ---
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name='report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5001, host='0.0.0.0')